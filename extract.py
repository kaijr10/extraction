"""
Instruction to run the script:
1. Required Python 3
2. Required pip3:
    apt install python3-pip
3. Install pdfminer for reading PDF file
    pip3 install pdfminer.six
4. Install pandas for write xml file
    pip3 install openpyxl xlsxwriter xlrd
    pip3 install pandas
5. Run the script:
    python3 extract.py <folder containing pdf files to extract> <output_excel_file>
    eg.
    python3 extract.py "folder contain pdf files" test.xlsx
"""


from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage
from io import StringIO
from openpyxl import load_workbook

import pandas as pd
import re
import glob
import sys
import os
import argparse


BESTELLING = "Bestelling:"
DATUM_BESTELLING = "Datum bestelling:"
REF = "ref:"
TOTAL = "Totaalbedrag Exc BTW:"
EURO_SIGN = "€"

row = 0

def convert_pdf_to_txt(path):
    rsrcmgr = PDFResourceManager()
    retstr = StringIO()
    codec = 'utf-8'

    """
    laparams=LAParams(all_texts=True, detect_vertical=True, 
                      line_overlap=0.5, char_margin=1000.0, #set char_margin to a large number
                      line_margin=0.5, word_margin=5,
                      boxes_flow=1)
    """
    laparams=LAParams(char_margin=10000.0, line_margin=10)
    device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
    fp = open(path, 'rb')
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    password = ""
    maxpages = 0
    caching = True
    pagenos=set()

    for page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages, password=password,caching=caching, check_extractable=True):
        interpreter.process_page(page)

    text = retstr.getvalue()

    fp.close()
    device.close()
    retstr.close()
    return text

class InvoiceProcessor(object):

    def __init__(self, text):
        self._text = text
        self._dataframe = {}
        self._result = []
        self._file_pass = True
    

    def process_text(self):
        # remove empty lines
        self._text = "\n".join([line for line in self._text.split("\n") if line.strip() != ''])
        # for each line
        result = {}
        index = 0
        temp_result = []
        for line in self._text.split("\n"):
            # merge multiple space to one space
            pattern = re.compile(r'\s+')
            line = re.sub(pattern, ' ', line)
            # handle for each row
            if "Datum bestelling:" in line and "ref:" in line:
                index = index + 1
                # split to get each item for bestelling, datum bestelling, ref_val
                bestelling, datatum_bestelling, ref = self.handle_for_bestelling(line)
                result[index] = []
                result[index].append(bestelling)
                result[index].append(datatum_bestelling)
                result[index].append(ref)
                # if temp_result is not None:
                #     result[index].append(temp_result)
                #     # reset temp_result
                #     temp_result = []
            elif "Totaalbedrag Exc BTW" in line:
                # handle for totaalbedragexcbtw
                totaal = self.handle_for_totaal(line)
                if (totaal is None and self._file_pass == False) or (totaal is not None and self._file_pass == True):
                    if totaal is None:
                        totaal = "N/A"
                    result[index].append(totaal)
                    if temp_result is not None:
                        result[index].append(temp_result)
                        # reset temp_result
                        temp_result = []
            else:
                if 'x' in line and '%' in line and EURO_SIGN in line:
                    line_item, dimension_1, dimension_2, bwt, prijs, totaal = self.handle_for_line_item_dimesions_so_on(line)
                    temp_result.append([line_item, dimension_1, dimension_2, bwt, prijs, totaal])

        return result


    def extract(self):
        result = self.process_text()
        f_data = []
        for _, value in result.items():
            data = []
            for ele in value[:-1]:
                data.append(ele)
            for ele in value[-1]:
                f_data.append(data + ele)
        return f_data
                
    def handle_for_bestelling(self, text):
        """
        example line: 
        Good: Bestelling: 264306, Datum bestelling: 2021-05-10, ref: B 100068766 Lucian Simo
        Bad: Datum bestelling: 2021-06-11 (Bestelling: 270863, ref: PCB 100070747 Pascal Janssen)    Datum of Aflever: 2021-06-28 (Levering GLS tracking #:646830032560)

        Return (264306, 2021-05-10, B 100068766)
        """
        if "(Bestelling:" in text:
            self._file_pass = False
            text = text.split(')')[0].replace('(', '')
            bestelling = text[(text.find("Bestelling:") + len("Bestelling:")):text.find("ref:")]
            datum_bestelling = text[(text.find("Datum bestelling:") + len("Datum bestelling:")):text.find("Bestelling:")]
        else:
            bestelling = text[(text.find("Bestelling:") + len("Bestelling:")):text.find("Datum bestelling:")]
            datum_bestelling = text[(text.find("Datum bestelling:") + len("Datum bestelling:")):text.find("ref:")]
        ref = text.split("ref:")[1]

        return (self.remove_special_characters(bestelling), self.remove_special_characters(datum_bestelling), self.remove_special_characters(ref))
    
    def handle_for_totaal(self, text):
        try:
            return self.remove_special_characters(text.split("Totaalbedrag Exc BTW:")[1])
        except IndexError as e:
            return None
    
    def handle_for_line_item_dimesions_so_on(self, text):
        """
        Example line
        1 x houten jaloezieën 50mm 0.00% € 93.04  € 93.04[1160 x 1770 mm]
        """
        # split "%""
        items = text.split("%")
        line_item = " ".join(items[0].split(" ")[:-1])
        bwt = " ".join([items[0].split(" ")[-1],"%"])
        child_item = [ele.strip() for ele in items[1].strip().split(EURO_SIGN) if ele != '']
        prijs = child_item[0]
        if '[' in text and ']' in text:
            dimension_1 = child_item[1].split('[')[1].split("x")[0].strip()
            dimension_2 = child_item[1].split('[')[1].split("x")[1].strip().replace(']', '')
            totaal = child_item[1].split('[')[0]
        else:
            dimension_1 = 0
            dimension_2 = 0
            totaal = child_item[1]

        return (line_item, dimension_1, dimension_2, bwt, prijs, totaal)

    
    @staticmethod
    def remove_special_characters(text):
        pattern = r'[:|,|€]'
        return re.sub(pattern, '', text).strip()

    def build_dataframe(self):

        result = self.extract()
        global row
        # handle result
        self._dataframe['Bestelling'] = []
        self._dataframe['Datum bestelling'] = []
        self._dataframe['ref'] = []
        self._dataframe['totaalbedrag exc BTW'] = []
        self._dataframe['line item'] = []
        self._dataframe['dimension 1 (mm)'] = []
        self._dataframe['dimension 2 (mm)'] = []
        self._dataframe['btw'] = []
        self._dataframe['prijs (ex)'] = []
        self._dataframe['totaal (ex)'] = []
        for ele in result:
            self._dataframe['Bestelling'].append(ele[0])
            self._dataframe['Datum bestelling'].append(ele[1])
            self._dataframe['ref'].append(ele[2])
            self._dataframe['totaalbedrag exc BTW'].append(ele[3])
            self._dataframe['line item'].append(ele[4])
            self._dataframe['dimension 1 (mm)'].append(ele[5])
            self._dataframe['dimension 2 (mm)'].append(ele[6])
            self._dataframe['btw'].append(ele[7])
            self._dataframe['prijs (ex)'].append(ele[8])
            self._dataframe['totaal (ex)'].append(ele[9])

        return self._dataframe



class AnwisProcessor(object):
    CATALOGUE = "Catalogue no. Order No"
    ITEM_NAME = "Item Name"
    DIMESION_1 = "Dimension 1"
    DIMESION_2 = "Dimension 2"
    ITEM_NO = "Item No"
    REF = "ref"
    CUSTOMER = "Customer"
    QUANTITY = "Quantity"
    NET_PRICE = "Net Price"
    NET_PRICE_AFTER_DISCOUNT = "Net Price After Discount"
    NET_AMOUNT = "Net Amount After Discount"
    GROSS_AMOUNT = "Grossamount After Discount"

    def __init__(self):
        pass

    def read_pdf_file(self, pdf_file):
        tables = read_pdf(pdf_file, pages='all', encoding='utf-8')
        return tables

    def main(self, pdf_file):
        result = {}
        tables = self.read_pdf_file(pdf_file)
        for table in tables:
            # convert table as dataframe to json
            data = self.convert_table_dataframe_to_json(table)
            # extract for each table
            data_result = self.extract_data(data)
            #print(data_result)
            if len(data_result[self.ITEM_NAME]) != 0:
                if not result:
                    result = data_result
                else:
                    for key in result.keys():
                        result[key] = result[key] + data_result[key]
        return result



    def extract_data(self, data):
        key_int = 0
        result = {}
        result[self.ITEM_NAME] = []
        result[self.DIMESION_1] = []
        result[self.DIMESION_2] = []
        result[self.ITEM_NO] = []
        result[self.REF] = []
        result[self.CUSTOMER] = []
        result[self.QUANTITY] = []
        result[self.NET_PRICE] = []
        result[self.NET_PRICE_AFTER_DISCOUNT] = []
        result[self.NET_AMOUNT] = []
        result[self.GROSS_AMOUNT] = []
        #print(json.dumps(data, indent=4))
        #sys.exit(0)
        for key, value in data.items():
            for key_c, value_c in value.items():
                if key_c == "Net" and value_c is not None and value_c.replace(",", "").isdigit():
                    #print(key, value)
                    item_dimensions = data[str(key_int + 1)][self.CATALOGUE]
                    item_no_ref_customer = data[str(key_int + 2)][self.CATALOGUE]
                    # process_data
                    item_name, dimension_1, dimension_2 = self.process_item_dimensions(item_dimensions)
                    item_no, ref, customer = self.process_item_no_ref_customer(item_no_ref_customer)
                    try:
                        quantity = value["Unnamed: 2"]
                        net_price = value["Unnamed: 4"]
                        net_price_afer_discount = value["Net"]
                        net_mount = value["Net.1"]
                        if 'Gross' in value.keys():
                            gross_amount = value["Gross"]
                        else:
                            gross_amount = value["Unnamed: 8"]
                    except Exception as e:
                        quantity = value["Quantity"]
                        net_price = value["Net"]
                        net_price_afer_discount = value["Net  price"]
                        net_mount = value["Net amount"]
                        gross_amount = value["Grossamount after"]

                    result[self.ITEM_NAME].append(item_name)
                    result[self.DIMESION_1].append(dimension_1)
                    result[self.DIMESION_2].append(dimension_2)
                    result[self.ITEM_NO].append(item_no)
                    result[self.REF].append(ref)
                    result[self.CUSTOMER].append(customer)
                    result[self.QUANTITY].append(quantity)
                    result[self.NET_PRICE].append(net_price)
                    result[self.NET_PRICE_AFTER_DISCOUNT].append(net_price_afer_discount)
                    result[self.NET_AMOUNT].append(net_mount)
                    result[self.GROSS_AMOUNT].append(gross_amount)

            key_int = key_int + 1

        return result

    @staticmethod
    def process_item_dimensions(text):
        if "/" not in text:
            dimension_1 = 0
            dimension_2 = 0
            item_name = text
        else:
            values = text.split("/")
            dimension_1 = values[0].split(" ")[-1].strip()
            dimension_2 = values[1].split(" ")[0].strip()
            item_name = " ".join(values[0].split(" ")[:-1])
        
        return (item_name, dimension_1, dimension_2)

    @staticmethod
    def process_item_no_ref_customer(text):
        values = text.split(" ")
        item_no = values[0]
        if len(values) == 1:
            ref = "N/A"
            customer = "N/A"
        elif len(values) == 2:
            ref = "N/A"
            customer = values[1]
        else:
            if values[1].isdigit():
                ref = values[1]
                customer = " ".join(values[2:])
            elif values[2].isdigit() and len([val for val in values if val.strip() != '']) > 3:
                ref = " ".join([values[1], values[2]])
                customer = " ".join(values[2:])
            else:
                ref = "N/A"
                customer = " ".join(values[1:])
        
        return (item_no, ref, customer)

    @staticmethod
    def convert_table_dataframe_to_json(table):
        parsed = json.loads(table.to_json(orient="index"))
        #print(json.dumps(parsed, indent=4))
        return parsed

def process_text(text, type_extract="invoice"):
    if type_extract == "invoice":
        invoice_processor = InvoiceProcessor(text)
        invoice_processor.process_text()
        dataframe = invoice_processor.build_dataframe()
    elif type_extract == "anwis":
        result = anwis_extract(text)
        dataframe = anwis_build_dataframe(result)
    else:
        result = toppoint_extract(text)
        dataframe = toppoint_build_dataframe(result)

    return dataframe

def save_to_excel_file(data, output_file, truncate_sheet=False, sheet_name='Sheet1', startrow=None):
    # save to output file
    df = pd.DataFrame(data)
    if not os.path.isfile(output_file):
        df.to_excel(output_file, index=None, sheet_name='Sheet1', startrow=0)
    else:
        # append it
        writer = pd.ExcelWriter(output_file, engine='openpyxl', mode='a')

        # try to open an existing workbook
        writer.book = load_workbook(output_file)
        
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, index=False, header=None)

        # save the workbook
        writer.save()


def parser_agrs():
    # initiate parser
    parser = argparse.ArgumentParser(description="The script tool to extract pdf files into excel file report")

    # add arguments
    parser.add_argument('-i', '--input', type=str, required=True, help="The folder contains pdf files to be extracted")
    parser.add_argument('-o', '--output', type=str, required=True, help="The output file xlsx file")
    parser.add_argument('-t', '--type', type=str, required=True, help="The type format is in : {\"invoice\", \"anwis\", \"toppoint\"}")
    
    args = parser.parse_args()
    
    return args


def main(args):
    pdf_folder = args.input
    output_file = args.output
    if not os.path.isdir(pdf_folder):
        print("ERROR - foder {pdf_folder} not exist or incorrect!".format(pdf_folder=pdf_folder))
    
    for pdf_file in glob.glob(pdf_folder + "/*.pdf"):
        try:
            if args.type == "invoices":
                # read pdf file
                text = convert_pdf_to_txt(pdf_file)
                # process the data
                dataframe = process_text(text)
            elif args.type == "anwis":
                dataframe = AnwisProcessor().main(pdf_file)
            save_to_excel_file(dataframe, output_file)
            print("{pdf_file} completed without any issues".format(pdf_file=pdf_file))
        except Exception:
            print("{pdf_file} - have issued".format(pdf_file=pdf_file))
            pass


if __name__ == "__main__":
    main(parser_agrs())